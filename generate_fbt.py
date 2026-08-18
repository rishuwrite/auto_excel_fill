#!/usr/bin/env python3
"""
generate_fbt.py — Split a FedEx "booking" export into the two FBT batch
upload templates (US / Non-US), by filling YOUR ACTUAL template files
in place: same formatting, same dropdown lists, same everything — with
live Excel formulas, saved back out as true Excel 97-2003 (.xls).

How it works (and why):
    A plain .xls (BIFF8) file can't be edited directly in Python while
    keeping its formatting and dropdown (data validation) lists intact —
    no maintained Python library reads+writes BIFF8 with full fidelity.
    So this script:
      1. Converts your template .xls -> .xlsx with LibreOffice (lossless,
         keeps formatting + dropdowns).
      2. Edits the .xlsx with openpyxl (keeps formatting + dropdowns,
         writes real formulas).
      3. Converts the edited .xlsx back -> .xls with LibreOffice, which
         writes genuine BIFF8 formula records (not just cached values)
         and preserves the dropdown lists.

Requires:
    pip install openpyxl xlrd
    On Windows, Microsoft Excel is preferred automatically (no admin install needed
    if Excel is already installed); LibreOffice is used as a fallback.
    On Linux/macOS, LibreOffice is used.

Usage:
    python3 generate_fbt.py BOOKING_FILE.xls \
        --mode yes \
        --us-template FBT_US_NFEI_YES.xls \
        --nonus-template FBT_NON_US_NFEI_YES.xls \
        --outdir ./out
"""

import argparse
import copy
import datetime
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
import xlrd
import openpyxl


# ---------------------------------------------------------------------------
# Column index (1-based, openpyxl-style) for every field in
# "Recipient and Invoice Data". Same for US / Non-US templates — they only
# differ in the header text of columns 1 and 39 (SequenceNumber/QTY names),
# which we never touch.
# ---------------------------------------------------------------------------
COL = {
    'SequenceNumber': 1, 'Recipient_ContactName': 2, 'Recipient_CompanyName': 3,
    'Recipient_AddressLine1': 4, 'Recipient_AddressLine2': 5, 'Recipient_AddressLine3': 6,
    'Recipient_Country': 7, 'Recipient_City': 8, 'Recipient_State': 9, 'Recipient_Postalcode': 10,
    'Recipient_PhoneNumber': 11, 'Reference_1': 15, 'InvoiceNumber': 20, 'InvoiceDate': 21,
    'TotalNoofPackage': 22, 'TotalShipmentweight': 23, 'Freight_charges': 27,
    'Insurance_charges': 28, 'Other_charges': 29, 'FOBValue': 31, 'InvoiceValue': 33,
    'Currency': 34, 'CountryofManufacture': 35, 'Commodity': 36, 'HSCODE1': 37,
    'StofOriginofgoods': 38, 'DisOfOriginofgoods': 39, 'QTY': 40, 'UOM1': 41,
    'UNIT_VALUE1': 42, 'UNIT_Weight1': 43, 'AdditionalInfo': 46,
}
LAST_COL = 50  # AX

ADDITIONAL_INFO_TEXT = (
    'FDADeviceListing#D394769,Code-HQG(Lens,spectacle,Prescription\n'
    'Eyeglasses)LensKartisbothManufacturer&Shipper'
)


# ---------------------------------------------------------------------------
# Spreadsheet conversion helpers
# ---------------------------------------------------------------------------
def find_soffice():
    for name in ('soffice', 'libreoffice'):
        path = shutil.which(name)
        if path:
            return path
    return None


def _excel_convert(src, target_format, outdir):
    """Use installed Microsoft Excel on Windows when available.

    This is especially useful on managed office PCs: Excel is often already
    installed, and using it avoids requiring administrator rights to install
    LibreOffice.
    """
    if sys.platform != 'win32':
        return None
    try:
        import win32com.client
    except ImportError:
        return None

    src = Path(src).resolve()
    outdir = Path(outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    ext = target_format.split(':')[0].lower()
    fmt = 51 if ext == 'xlsx' else 56 if ext == 'xls' else None
    if fmt is None:
        return None

    # Always create the target with a distinct filename in outdir.
    target = outdir / (src.stem + '.' + ext)
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(src), UpdateLinks=0, ReadOnly=False)
        wb.SaveAs(str(target), FileFormat=fmt)
        wb.Close(SaveChanges=False)
        wb = None
        if not target.exists():
            raise RuntimeError(f"Excel did not create expected output: {target}")
        return target
    except Exception:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        return None
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def convert(soffice, src, target_format, outdir):
    """Convert xls/xlsx using Excel on Windows, otherwise LibreOffice."""
    excel_result = _excel_convert(src, target_format, outdir)
    if excel_result is not None:
        return excel_result

    if not soffice:
        sys.exit(
            'ERROR: No spreadsheet converter is available. On Windows, install '
            'Microsoft Excel or LibreOffice. On Linux/macOS, install LibreOffice.'
        )

    cmd = [soffice, '--headless', '--convert-to', target_format,
           '--outdir', str(outdir), str(src)]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        sys.exit(f'ERROR converting {src} -> {target_format}:\n{result.stdout}\n{result.stderr}')
    produced = Path(outdir) / (Path(src).stem + '.' + target_format.split(':')[0])
    if not produced.exists():
        sys.exit(f'ERROR: expected output {produced} was not created.\n{result.stdout}')
    return produced


# ---------------------------------------------------------------------------
# Booking data parsing (xlrd, same rules as before)
# ---------------------------------------------------------------------------
def normalize(h):
    return re.sub(r'[^a-z0-9]', '', str(h or '').lower())


FIELD_ALIASES = {
    'refnumber': ['refnumber'], 'invoice': ['invoice'], 'custname': ['custname'],
    'custaddress': ['custaddress'], 'custcity': ['custcity'], 'state': ['state'],
    'pin': ['pin'], 'custmobile': ['custmobile'], 'content': ['content'], 'qty': ['qty'],
    'destinationcountry': ['destinationcountry'],
}


def find_booking_sheet(wb):
    for name in wb.sheet_names():
        if normalize(name) == 'booking':
            return name
    for name in wb.sheet_names():
        if 'booking' in normalize(name):
            return name
    return None


def cell_str(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def cell_num(v):
    if v in (None, ''):
        return 0.0
    try:
        return float(str(v).replace(',', ''))
    except ValueError:
        return 0.0


def split_address(addr):
    text = (addr or '').strip()
    if len(text) > 105:
        return [text, '', '']
    words = text.split()
    lines = ['', '', '']
    li = 0
    for w in words:
        if li >= 3:
            lines[2] = (lines[2] + ' ' + w).strip()
            continue
        candidate = (lines[li] + ' ' + w).strip() if lines[li] else w
        if len(candidate) <= 35:
            lines[li] = candidate
        else:
            li += 1
            if li >= 3:
                lines[2] = (lines[2] + ' ' + w).strip()
            else:
                lines[li] = w
    return lines


def excel_today_serial():
    """Return today's date as an Excel 1900-system serial number."""
    today = datetime.date.today()
    epoch = datetime.date(1899, 12, 31)
    serial = (today - epoch).days
    # Excel incorrectly treats 1900 as a leap year, so serials from
    # 1900-03-01 onward are one higher than the simple day difference.
    if today >= datetime.date(1900, 3, 1):
        serial += 1
    return serial

def today_ddmmyy():
    # Kept for the output filename/date label only.
    return datetime.date.today().strftime('%d%m%y')


def hs_code(commodity, nfei_yes, is_us):
    c = commodity or ''
    is_eyeglasses = re.search(r'Prescription\s*Eyeglasses', c, re.I)
    is_sunglasses = re.search(r'Polarized\s*Sunglasses', c, re.I)
    short_codes = nfei_yes and not is_us
    if is_eyeglasses:
        return '90049090' if short_codes else '9004900090'
    if is_sunglasses:
        return '90041000' if short_codes else '9004100000'
    return ''


def parse_booking_rows(sheet):
    header_row = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    col_map = {}
    for idx, h in enumerate(header_row):
        n = normalize(h)
        if n and n not in col_map:
            col_map[n] = idx

    idx = {}
    for key, aliases in FIELD_ALIASES.items():
        idx[key] = next((col_map[a] for a in aliases if a in col_map), -1)

    idx['netdeclaredamount'] = -1
    for i, h in enumerate(header_row):
        hs = str(h or '').lower()
        if 'net' in hs and 'decl' in hs and 'amount' in hs:
            idx['netdeclaredamount'] = i
            break

    missing = [k for k, v in idx.items() if v == -1]
    if missing:
        raise ValueError(f'Could not find required column(s) in booking sheet: {", ".join(missing)}')

    records = []
    for r in range(1, sheet.nrows):
        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if all(v in (None, '') for v in row):
            continue
        records.append({
            'refNumber': cell_str(row[idx['refnumber']]),
            'invoice': cell_str(row[idx['invoice']]),
            'custName': cell_str(row[idx['custname']]),
            'custAddress': cell_str(row[idx['custaddress']]),
            'custCity': cell_str(row[idx['custcity']]),
            'state': cell_str(row[idx['state']]),
            'pin': cell_str(row[idx['pin']]),
            'custMobile': cell_str(row[idx['custmobile']]),
            'content': cell_str(row[idx['content']]),
            'qty': cell_num(row[idx['qty']]),
            'netDeclaredAmount': cell_num(row[idx['netdeclaredamount']]),
            'destinationCountry': cell_str(row[idx['destinationcountry']]),
        })
    return records


# ---------------------------------------------------------------------------
# Template editing (openpyxl)
# ---------------------------------------------------------------------------
def find_sheet(wb, wanted_norm):
    for name in wb.sheetnames:
        if normalize(name) == wanted_norm:
            return wb[name]
    for name in wb.sheetnames:
        if wanted_norm in normalize(name):
            return wb[name]
    return None


def clear_data_rows(ws, from_row, to_row, last_col):
    for r in range(from_row, to_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).value = None


def extend_validation_ranges(ws, new_last_row):
    """Stretch every dropdown/list validation on this sheet so it still
    covers every data row we wrote, even if we wrote more rows than the
    template originally had."""
    from openpyxl.utils import get_column_letter, range_boundaries
    for dv in list(ws.data_validations.dataValidation):
        new_sqref_parts = []
        for rng in str(dv.sqref).split():
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            if min_row and min_row > 1:  # skip header-row-only validations
                max_row = new_last_row
                new_sqref_parts.append(
                    f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}'
                )
            else:
                new_sqref_parts.append(rng)
        dv.sqref = ' '.join(new_sqref_parts)


def last_value_row_col(ws):
    """Return the last row/column containing an actual value or formula."""
    last_row = 0
    last_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ''):
                last_row = max(last_row, cell.row)
                last_col = max(last_col, cell.column)
    return last_row, last_col


def trim_trailing_rows_and_columns(ws, min_data_row=2):
    """Physically remove trailing blank rows/columns after the last field.

    This is intentionally based on values/formulas, not formatting. It keeps
    the workbook compact while retaining the populated fields.
    """
    last_row, last_col = last_value_row_col(ws)
    if last_row < min_data_row:
        last_row = min_data_row - 1
    if last_col < 1:
        last_col = 1

    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)
    if ws.max_column > last_col:
        ws.delete_cols(last_col + 1, ws.max_column - last_col)


def copy_data_row_format(ws, source_row, target_row):
    """Copy the template data-row formatting to target_row."""
    if target_row == source_row:
        return
    for c in range(1, LAST_COL + 1):
        src = ws.cell(row=source_row, column=c)
        dst = ws.cell(row=target_row, column=c)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        if src.protection:
            dst.protection = copy.copy(src.protection)
    src_dim = ws.row_dimensions[source_row]
    dst_dim = ws.row_dimensions[target_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden


def fill_recipient_sheet(ws, records, nfei_yes, is_us):
    """Formula-only mode.

    Rules:
      - No Python calculation or rounding of monetary amounts.
      - InvoiceValue comes directly from the booking export and is NOT
        overwritten by a template formula.
      - Every formula already present in template row 2 is dragged down with
        relative references translated to each generated row.
      - UNIT_VALUE1 is NOT replaced with ROUND/TRUNC; the template's own
        formula is used exactly as supplied.
      - Date is a real Excel 1900-system serial and is displayed as ddmmyy.
      - The complete row-2 formatting/borders are copied to every generated row.
    """
    original_last_row = ws.max_row

    # IMPORTANT: capture the template's row-2 formulas BEFORE clearing any
    # data rows. This is the source of truth for every formula in the output.
    # InvoiceValue is deliberately excluded: its amount must remain exactly
    # the booking amount, with no Python/template recalculation.
    template_formulas = {}
    for c in range(1, LAST_COL + 1):
        if c == COL['InvoiceValue']:
            continue
        value = ws.cell(row=2, column=c).value
        if isinstance(value, str) and value.startswith('='):
            template_formulas[c] = value

    # Now clear old data rows.
    clear_data_rows(ws, 2, original_last_row, LAST_COL)

    date_serial = excel_today_serial()

    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter

    for i, rec in enumerate(records):
        r = i + 2

        # Copy the complete template data-row style first, including borders,
        # number formats, protection and row height.
        copy_data_row_format(ws, 2, r)

        addr1, addr2, addr3 = split_address(rec['custAddress'])

        values = {
            'SequenceNumber': i + 1,
            'Recipient_ContactName': rec['custName'],
            'Recipient_CompanyName': rec['custName'],
            'Recipient_AddressLine1': addr1,
            'Recipient_AddressLine2': addr2,
            'Recipient_AddressLine3': addr3,
            'Recipient_Country': rec['destinationCountry'],
            'Recipient_City': rec['custCity'],
            'Recipient_State': rec['state'],
            'Recipient_Postalcode': rec['pin'],
            'Recipient_PhoneNumber': rec['custMobile'],
            'Reference_1': rec['refNumber'],
            'InvoiceNumber': rec['invoice'],

            # Real Excel date serial. The number format below makes Excel
            # display it as ddmmyy, e.g. 180826.
            'InvoiceDate': date_serial,

            'TotalNoofPackage': 1,
            'TotalShipmentweight': 0.5,
            'Currency': 'US DOLLARS-USD',
            'CountryofManufacture': 'IN-INDIA',
            'Commodity': rec['content'],
            'HSCODE1': hs_code(rec['content'], nfei_yes, is_us),
            'StofOriginofgoods': 'Gurugram',
            'DisOfOriginofgoods': 'Haryana',
            'QTY': rec['qty'],
            'UOM1': 'PIECE',
            'AdditionalInfo': ADDITIONAL_INFO_TEXT,
        }

        # EXACT booking InvoiceValue. No calculation, rounding, truncation,
        # adjustment or formula replacement.
        values['InvoiceValue'] = rec['netDeclaredAmount']

        for key, value in values.items():
            ws.cell(row=r, column=COL[key]).value = value

        # The only fixed US values required outside the template formulas.
        # These are constants, not calculations.
        if is_us:
            ws.cell(row=r, column=COL['Freight_charges']).value = 13.5
            ws.cell(row=r, column=COL['Insurance_charges']).value = 0.5
        else:
            ws.cell(row=r, column=COL['Freight_charges']).value = None
            ws.cell(row=r, column=COL['Insurance_charges']).value = None

        # Drag ONLY the formulas that already exist in template row 2.
        # This includes UNIT_VALUE1, FOBValue, Other_charges, etc., exactly
        # as the template defines them. No new ROUND/TRUNC formula is added.
        for col, formula in template_formulas.items():
            origin = f'{get_column_letter(col)}2'
            target = f'{get_column_letter(col)}{r}'
            try:
                translated = Translator(formula, origin=origin).translate_formula(target)
            except Exception:
                translated = formula
            ws.cell(row=r, column=col).value = translated

        # Force the required date display without changing the underlying
        # Excel serial number.
        ws.cell(row=r, column=COL['InvoiceDate']).number_format = 'ddmmyy'

    new_last_row = len(records) + 1
    extend_validation_ranges(ws, new_last_row)

    # Reapply the complete template row formatting after all writes so the
    # bottom borders and formatting remain identical on every generated row.
    for r in range(2, new_last_row + 1):
        copy_data_row_format(ws, 2, r)
        ws.cell(row=r, column=COL['InvoiceDate']).number_format = 'ddmmyy'

    trim_trailing_rows_and_columns(ws, min_data_row=2)


def fill_importer_blank(ws):
    if ws is None:
        return
    clear_data_rows(ws, 2, max(ws.max_row, 2), ws.max_column or 11)
    trim_trailing_rows_and_columns(ws, min_data_row=2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def process_one(soffice, booking_records, template_xls, nfei_yes, is_us, outdir, mode_label, stamp, tmp):
    src_xlsx = convert(soffice, template_xls, 'xlsx', tmp)
    wb = openpyxl.load_workbook(src_xlsx)

    ws_recipient = find_sheet(wb, 'recipientandinvoicedata')
    if ws_recipient is None:
        sys.exit(f'ERROR: "Recipient and Invoice Data" sheet not found in {template_xls}')
    fill_recipient_sheet(ws_recipient, booking_records, nfei_yes, is_us)

    ws_importer = find_sheet(wb, 'importerinfo')
    fill_importer_blank(ws_importer)

    edited_xlsx = tmp / (Path(template_xls).stem + '_edited.xlsx')
    wb.save(edited_xlsx)

    region = 'US' if is_us else 'NON_US'
    out_name = f'FBT_{region}_NFEI_{mode_label}_{stamp}.xls'
    produced = convert(soffice, edited_xlsx, 'xls:MS Excel 97', outdir)
    final_path = outdir / out_name
    produced.replace(final_path)
    return final_path


def main():
    p = argparse.ArgumentParser(description='Fill your actual FBT US/Non-US templates from a booking export.')
    p.add_argument('booking_file', help='Path to the booking .xls export')
    p.add_argument('--mode', choices=['yes', 'no'], required=True, help='NFEI YES or NFEI NO rule set')
    p.add_argument('--us-template', required=True, help='Path to your real FBT_US_NFEI_*.xls template')
    p.add_argument('--nonus-template', required=True, help='Path to your real FBT_NON_US_NFEI_*.xls template')
    p.add_argument('--outdir', default='.', help='Output directory (default: current directory)')
    args = p.parse_args()

    soffice = find_soffice()
    nfei_yes = args.mode == 'yes'
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    wb_in = xlrd.open_workbook(args.booking_file)
    sheet_name = find_booking_sheet(wb_in)
    if not sheet_name:
        sys.exit(f'ERROR: no sheet named "booking" found. Sheets present: {wb_in.sheet_names()}')
    sheet = wb_in.sheet_by_name(sheet_name)
    print(f'Using sheet "{sheet_name}"')

    records = parse_booking_rows(sheet)
    print(f'Parsed {len(records)} booking row(s)')

    us = [r for r in records if r['destinationCountry'].strip().upper() == 'US']
    non_us = [r for r in records if r['destinationCountry'].strip().upper() != 'US']
    print(f'  -> {len(us)} US row(s)')
    print(f'  -> {len(non_us)} non-US row(s)')

    stamp = datetime.date.today().strftime('%Y-%m-%d')
    mode_label = 'YES' if nfei_yes else 'NO'

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        if us:
            out = process_one(soffice, us, args.us_template, nfei_yes, True, outdir, mode_label, stamp, tmp)
            print(f'Wrote {out} ({len(us)} rows)')
        else:
            print('No US rows - US file skipped.')

        if non_us:
            out = process_one(soffice, non_us, args.nonus_template, nfei_yes, False, outdir, mode_label, stamp, tmp)
            print(f'Wrote {out} ({len(non_us)} rows)')
        else:
            print('No non-US rows - non-US file skipped.')


if __name__ == '__main__':
    main()
